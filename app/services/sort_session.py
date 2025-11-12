"""Utilities for tracking sort sessions and exporting run data.

This module keeps an in-memory representation of the active sort session and
persists rows to an Excel workbook (one worksheet per session). When the Excel
writer backend is unavailable it falls back to CSV files so that data is not
lost during development environments.
"""
from __future__ import annotations

import asyncio
import json
import logging
import threading
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    import pandas as pd  # type: ignore[import-not-found]
except Exception:  # pragma: no cover - optional dependency during dev
    pd = None  # type: ignore[assignment]

LOG = logging.getLogger("sort.session")


def _sanitize_sheet_name(name: str) -> str:
    """Return a worksheet-safe name (Excel limits to 31 characters)."""
    invalid = set(':\\/*?[]')
    cleaned = ''.join(ch for ch in name if ch not in invalid)
    cleaned = cleaned.strip() or "Run"
    return cleaned[:31]


@dataclass
class SortSession:
    id: str
    sheet_name: str
    started_at: datetime
    meta: Dict[str, Any] = field(default_factory=dict)
    rows: List[Dict[str, Any]] = field(default_factory=list)
    last_entry: Optional[Dict[str, Any]] = None


class SortSessionManager:
    def __init__(self, workbook_path: str = "data/sort_runs.xlsx") -> None:
        self.workbook_path = Path(workbook_path)
        self._lock = threading.RLock()
        self._active: Optional[SortSession] = None

    # ------------------------------------------------------------------
    # lifecycle
    # ------------------------------------------------------------------
    def start_session(self, metadata: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """Begin a new session and create a fresh worksheet for it."""
        with self._lock:
            if self._active is not None:
                raise RuntimeError("sort session already active")

            meta = dict(metadata or {})
            now = datetime.utcnow()
            session_id = str(meta.get("session_id") or now.strftime("%Y%m%d-%H%M%S"))
            suggested = meta.get("sheet_name") or meta.get("session_name") or f"Run_{now.strftime('%m%d_%H%M')}"
            sheet_name = _sanitize_sheet_name(str(suggested))

            meta.update(
                {
                    "session_id": session_id,
                    "sheet_name": sheet_name,
                    "started_at": now.isoformat(timespec="seconds"),
                    "state": meta.get("state", "Running"),
                }
            )

            session = SortSession(id=session_id, sheet_name=sheet_name, started_at=now, meta=meta)
            self._active = session
            self._write_meta()
            LOG.info("Started sort session %s (sheet=%s)", session_id, sheet_name)
            return self._snapshot(session)

    def end_session(self, notes: Optional[str] = None) -> Dict[str, Any]:
        """Finalize the active session and persist metadata."""
        with self._lock:
            if self._active is None:
                raise RuntimeError("no active sort session")
            now = datetime.utcnow()
            if notes:
                self._active.meta["notes"] = notes
            self._active.meta["ended_at"] = now.isoformat(timespec="seconds")
            self._active.meta["state"] = "Complete"
            self._write_meta()
            snapshot = self._snapshot(self._active)
            LOG.info("Ended sort session %s", self._active.id)
            self._active = None
            return snapshot

    def update_state(self, state: str) -> Dict[str, Any]:
        """Update the session state (e.g. Running/Paused)."""
        with self._lock:
            if self._active is None:
                raise RuntimeError("no active sort session")
            self._active.meta["state"] = state
            self._write_meta()
            return self._snapshot(self._active)

    def ensure_session(self) -> SortSession:
        """Keep logging resilient by auto-starting a session if missing."""
        with self._lock:
            if self._active is None:
                LOG.info("Auto-starting sort session for logging")
                self.start_session({"session_name": "Auto"})
            assert self._active is not None
            return self._active

    # ------------------------------------------------------------------
    # recording
    # ------------------------------------------------------------------
    def record_operation(self, entry: Dict[str, Any]) -> Dict[str, Any]:
        with self._lock:
            session = self.ensure_session()
            row = dict(entry or {})
            row.setdefault("timestamp", datetime.utcnow().isoformat(timespec="seconds"))
            row.setdefault("sequence", len(session.rows) + 1)

            orientation = row.pop("orientation", None)
            if isinstance(orientation, dict):
                row["orientation_summary"] = orientation.get("determination")
                row["orientation_method"] = orientation.get("method")
                row["orientation_top_density"] = orientation.get("top_full_density")
                row["orientation_bottom_density"] = orientation.get("bottom_full_density")
                row["orientation_details"] = json.dumps(orientation, default=str)

            image_paths = row.pop("image_paths", None)
            if isinstance(image_paths, dict):
                row["image_top_path"] = image_paths.get("top")
                row["image_bottom_path"] = image_paths.get("bottom")

            session.rows.append(row)
            session.last_entry = row
            self._write_rows()
            return row

    # ------------------------------------------------------------------
    # status helpers
    # ------------------------------------------------------------------
    def status(self) -> Dict[str, Any]:
        with self._lock:
            if self._active is None:
                return {
                    "state": "Idle",
                    "total": 0,
                    "completed": 0,
                    "good": 0,
                    "err": 0,
                    "throughput_cpm": 0.0,
                    "progress_pct": 0,
                    "current_card": None,
                    "errors": [],
                }

            session = self._active
            count = len(session.rows)
            elapsed = max((datetime.utcnow() - session.started_at).total_seconds(), 1.0)
            cpm = (count / elapsed) * 60.0
            last = session.last_entry or {}
            return {
                "state": session.meta.get("state", "Running"),
                "session_id": session.id,
                "started_at": session.started_at.isoformat(timespec="seconds"),
                "total": count,
                "completed": count,
                "good": count,
                "err": 0,
                "throughput_cpm": round(cpm, 2),
                "progress_pct": min(100, count),
                "current_card": last.get("card_name"),
                "last_entry": last,
                "errors": [],
            }

    # ------------------------------------------------------------------
    # persistence helpers
    # ------------------------------------------------------------------
    def _write_rows(self) -> None:
        session = self._active
        if not session:
            return
        self._write_sheet(session.sheet_name, session.rows)

    def _write_meta(self) -> None:
        session = self._active
        if not session:
            return
        self._write_sheet(f"{session.sheet_name}_meta", [session.meta])

    def _write_sheet(self, sheet_name: str, rows: List[Dict[str, Any]]) -> None:
        safe_name = _sanitize_sheet_name(sheet_name)
        path = self.workbook_path
        path.parent.mkdir(parents=True, exist_ok=True)
        mode = "a" if path.exists() else "w"

        if pd is None:
            csv_path = path.with_name(f"{path.stem}_{safe_name}.csv")
            LOG.warning("pandas unavailable; logging rows to %s", csv_path)
            import csv

            with open(csv_path, "w", encoding="utf8", newline="") as fh:
                writer = csv.DictWriter(fh, fieldnames=list(rows[0].keys()) if rows else [])
                writer.writeheader()
                writer.writerows(rows)
            return

        df = pd.DataFrame(rows)
        if df.empty:
            df = pd.DataFrame([{}]).iloc[0:0]

        try:
            with pd.ExcelWriter(path, engine="openpyxl", mode=mode, if_sheet_exists="replace") as writer:
                df.to_excel(writer, sheet_name=safe_name, index=False)
        except TypeError:
            try:
                from openpyxl import load_workbook  # type: ignore[import-not-found]
            except Exception as exc:
                LOG.warning("openpyxl unavailable (%s); falling back to CSV for %s", exc, safe_name)
                csv_path = path.with_name(f"{path.stem}_{safe_name}.csv")
                df.to_csv(csv_path, index=False)
                return

            if path.exists() and mode == "a":
                book = load_workbook(path)
                if safe_name in book.sheetnames:
                    ws = book[safe_name]
                    book.remove(ws)
                    book.save(path)
            with pd.ExcelWriter(path, engine="openpyxl", mode=mode) as writer:
                df.to_excel(writer, sheet_name=safe_name, index=False)
        except (ImportError, ValueError) as exc:
            LOG.warning("Excel writer unavailable (%s); writing CSV fallback for %s", exc, safe_name)
            csv_path = path.with_name(f"{path.stem}_{safe_name}.csv")
            df.to_csv(csv_path, index=False)

    def _snapshot(self, session: SortSession) -> Dict[str, Any]:
        return {
            "session_id": session.id,
            "sheet_name": session.sheet_name,
            "started_at": session.started_at.isoformat(timespec="seconds"),
            "meta": dict(session.meta),
            "entries": len(session.rows),
            "last_entry": dict(session.last_entry) if session.last_entry else None,
        }


_manager = SortSessionManager()


def get_manager() -> SortSessionManager:
    return _manager


def record_operation(entry: Dict[str, Any]) -> Dict[str, Any]:
    return _manager.record_operation(entry)


async def log_operation_async(entry: Dict[str, Any]) -> Dict[str, Any]:
    loop = asyncio.get_running_loop()
    return await loop.run_in_executor(None, record_operation, entry)
